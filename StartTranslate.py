import os
import re
import sys
import copy
import json
from openpyxl import load_workbook

WORKING_DIR = os.path.dirname(os.path.realpath(__file__))

IN_FILE = os.path.join(WORKING_DIR, "TranslateData.xlsx")
OUT_FILE = os.path.join(WORKING_DIR, "SuperNewRoles", "Resources", "TranslateFile.json")

def stringToJson(in_files):
  stringData = {}
  for filename in in_files:
    if not os.path.isfile(filename):
      continue

    wb = load_workbook(filename, read_only = True)

    for s in wb:
      rows = s.iter_rows(min_col = 1, min_row = 2, max_col = 17, max_row = None)
      headers = []
      for header in s[1]:
        if header.value:
          headers.append(header.value)

      for row in rows:
        name = row[0].value

        if not name:
          continue

        data = {}

        for i, string in enumerate(row[1:]):
          if string.value:
            # I hate excel why did I do this to myself
            data[i] = string.value.replace("\r", "").replace("_x000D_", "").replace("\\n", "\n")

        if data:
          stringData[name] = data

  with open(OUT_FILE, "w", newline="\n") as f:
    json.dump(stringData, f, indent=4)

if __name__ == "__main__":
  in_files = [
    os.path.join(WORKING_DIR, "TranslateData.xlsx"),
    os.path.join(WORKING_DIR, "TranslateData-Dev.xlsx")
  ]

  stringToJson(in_files)